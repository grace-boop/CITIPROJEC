import pdfplumber
import nltk
import fitz  # PyMuPDF
import os
import re
import numpy as np
#import gensim.downloader as api
from nltk.stem import WordNetLemmatizer
from gensim.downloader import load
from gensim.models import KeyedVectors
from gensim.models import Word2Vec
from gensim.models.word2vec import LineSentence
from nltk.tokenize import word_tokenize
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from fuzzywuzzy import process

#Intialization
#nltk.download('punkt')
#nltk.download('wordnet')
#wordlist:關鍵字list，可以改!!!!!
wordlist = ["block-chain-related","cloud technology","data technology",
"internet technology","blockchain","cloud computing","big data",
"mobile","alliance chain","cloud architecture","data layer","internet",
"test chain","cloud service","dataset","network","interconnected chain",
"cloud finance","data flow","online"]
#不用訓練好的model因為沒有很多字
#以下適用api的model 不知道跟KeyVectored差在哪裡
'''
#word2vec_model = api.load("word2vec-google-news-300")
for word in wordlist:
    if word in word_vectors.key_to_index:
        print(f"The word '{word}' is in the pretrained model.")
    else:
        vocabulary = list(word2vec_model.index_to_key)
        similar_word, _ = process.extractOne(word, vocabulary)
        if similar_word:
            similar_words = word2vec_model.similar_by_word(word)
            similar_words = similar_words[:5]
        for similarword, similarity in similar_words:
            if similarword in word2vec_model.key_to_index:
                print(f"The word '{word}' is similar to'{similarword,similarity}' in the pretrained model.")
                break
            else:
                print(f"The word '{word}' is not in the pretrained model.")
'''
#stoplist:if出現這些字則句子不算數，可以改!!!!
stoplist = ["not","no"]
'''
def sentence_vector(seq, model):

    vectors = []
    for word in seq:
        #print(model.wv.vocab)
        if word in model.wv:
            print(model.wv[word])
            vectors.append(model.wv[word])
    if len(vectors) == 0:
        return np.zeros(model.vector_size)
    return np.mean(vectors, axis=0)
'''
def lemmatize_sentence(sentence):
    lemmatizer = WordNetLemmatizer()
    tokens = nltk.word_tokenize(sentence)
    lemmatized_tokens = [lemmatizer.lemmatize(token.lower()) for token in tokens]
    lemmatized_sentence = " ".join(lemmatized_tokens)
    return lemmatized_sentence

def citiproject(files,dir):
    workbook = Workbook()
    bankname = dir
    #file是一個pdf
    for file in files:
        print(file)
        with fitz.open('./data/'+dir+'/'+file) as pdf_document:
            #word_frequency = [0 for _ in range(len(wordlist))] 
            my_dict = {}
            #textLIst: one index one page 
            textlist = []
            #預處理的text:有多寫了不用的東西弄完要再改一下
            for page_number in range(pdf_document.page_count):
                page = pdf_document.load_page(page_number)
                text = page.get_text()
                textlist.append(text)

            '''
            process_text = []
            process_word = []
            for text in textlist:
                sentences = nltk.sent_tokenize(text)#沒有切詞，不知道會不會有差
                for sentence in sentences:
                    #word_in_sent = sentence.lower().count(wordlist[i])
                    #pattern = r"\b"+re.escape(wordlist[i])+r"\b"
                    lemmatized_sentence = lemmatize_sentence(sentence)
                    for word in wordlist:
                        lemmatized_sentence = re.sub(r'\b' + re.escape(word) + r'\b', word.replace(' ', '_'), lemmatized_sentence)
                    words = word_tokenize(lemmatized_sentence)
                    process_text.append(lemmatized_sentence)
                    process_word.append(words)
            print(process_text[0])
            print(process_text[1])
            print(process_word[0])
            print(process_word[1])
            
            for page_number in range(pdf_document.page_count):
                page = pdf_document.load_page(page_number)
                text = page.get_text()
                sentences = nltk.sent_tokenize(text)#沒有切詞，不知道會不會有差
                for sentence in sentences:
                    #word_in_sent = sentence.lower().count(wordlist[i])
                    #pattern = r"\b"+re.escape(wordlist[i])+r"\b"
                    lemmatized_sentence = lemmatize_sentence(sentence)
                    for word in wordlist:
                        lemmatized_sentence = re.sub(r'\b' + re.escape(word) + r'\b', word.replace(' ', '_'), lemmatized_sentence)
                    words = word_tokenize(lemmatized_sentence)
                    lemmatext.append(lemmatized_sentence)
                    process_word.append(words)
                textlist.append(text)
            #print(textlist)
            
            #存進txt
            with open('sentences.txt', 'w',encoding='utf-8') as f:
                for text in process_text:
                    f.write(text + '\n')
            
            # 使用 LineSentence 加载文本文件
            sentences = LineSentence('sentences.txt')
            model = Word2Vec(process_word, vector_size=100, window=5, min_count=5, workers=4)
            for i in range(len(model.wv.index_to_key)):
                for word in wordlist:
                    model.wv.index_to_key[i] = model.wv.index_to_key[i].replace(word.replace(' ', '_'), word)
            
            key_to_index = model.wv.key_to_index
            index_to_key = model.wv.index_to_key
            for key in wordlist:
                print(key)
                for word, index in key_to_index.items():
                    if key == word:##沒出現應該是因為機率太低
                        print(word, index)
            '''

            word_matrix = [[0 for _ in range(len(textlist))] for _ in range(len(wordlist))]
            # len(wordlist) rows, len(textlist) columns??
            for i in range(len(wordlist)):
                ##appear in which pages
                page_number = []
                ##appear in which sentences
                sen_list = []
                #word_vector = word2vec_model[wordlist[i]]
                for j in range(len(textlist)):
                    #appear in which sentences, an index is in the same page_number
                    seq=[]
                    sentences = nltk.sent_tokenize(textlist[j])#沒有切詞，不知道會不會有差
                    for sentence in sentences:
                        #word_in_sent = sentence.lower().count(wordlist[i])
                        pattern = r"\b"+re.escape(wordlist[i])+r"\b"
                        lemmatized_sentence = lemmatize_sentence(sentence)
                        word_in_sent = len(re.findall(pattern, lemmatized_sentence, re.IGNORECASE))
                        if word_in_sent>0:
                            #determine whether the sentence has stopwprds or not, stop_flag = 1 if its has stopword
                            stop_flag = 0
                            for stop in stoplist:
                                temp_sent = lemmatized_sentence.split()
                                if stop.lower() in [w.lower() for w in temp_sent]:
                                    stop_flag = 1
                            if stop_flag == 1:
                                continue
                            else:
                                word_matrix[i][j]+=word_in_sent
                                seq.append(sentence)###有換行符號 在看看要怎樣
                            
                    #the sentences append in 
                    if len(seq)>0:
                        sen_list.append(seq)
                        '''
                        vector = sentence_vector(seq, model)
                        print(seq)
                        print(vector)
                        '''
                    #the Ith word is appear in the Jth page
                    if word_matrix[i][j] > 0:
                        page_number.append(j) 
                temp_count = 0
                for j in range(len(textlist)):
                    temp_count+=word_matrix[i][j]
                new_row = {"frequence":temp_count,"page_number":page_number,"sentences":sen_list}
                my_dict.update({wordlist[i]:new_row})

            #print(my_dict)

            ##存到excel
            temp_sheet = workbook.create_sheet(title=file)
            '''
            #wb = Workbook()
            #ws = wb.active
            #ws.append(["word", "frequence", "page_number", "sentences"])
            '''
            temp_sheet.append(["word"])
            row_idx = 2
            for word, word_data in my_dict.items():

                temp_sheet.cell(row=row_idx, column=1, value=word)
                #if not word_data["sentences"]:
                #    continue  # 如果为空列表，则跳过当前行的写入

                for column_idx, (key, value) in enumerate(word_data.items(), start=2):
                    # 如果值是列表，则将列表中的每个元素写入到相应的列中
                    if isinstance(value, list):
                        if key == "page_number":
                            for idx, page_number in enumerate(value, start=1):
                                temp_sheet.cell(row=row_idx + idx -1, column=column_idx, value=page_number)
                        else:
                            for sublist_idx, sublist in enumerate(value, start=1):
                                for subvalue_idx, subvalue in enumerate(sublist, start=1):
                                    cell = temp_sheet.cell(row=row_idx + sublist_idx - 1, column=column_idx + subvalue_idx - 1, value=subvalue)

                                    column_letter = get_column_letter(cell.column)
                                    temp_sheet.column_dimensions[column_letter].width = max(temp_sheet.column_dimensions[column_letter].width, len(subvalue)+2)
                    else:
                        temp_sheet.cell(row=row_idx, column=column_idx, value=value)
                row_idx +=max(1, len(word_data["sentences"]))
    # 删除工作表
    sheet_to_delete = workbook['Sheet'] 
    workbook.remove(sheet_to_delete)
    # 保存 Excel 文件
    workbook.save("./result/"+dir+".xlsx")

def traverse_folders(root_path):
    for root, dirs, files in os.walk(root_path):
        for dir in dirs:
            print("資料夾:", dir)
            path = "./data/"+dir
            files = os.listdir(path)
            citiproject(files,dir)
            #!!!
            #for file in files:
            #    citiproject(file)
            #    print(file)
            #!!!

root_path = "./data/"
traverse_folders(root_path)


#一層directory的方法
'''
files = os.listdir(folder_path)

for file in files:
    #citiproject(file)
    print(file)
#pdf plumber
with pdfplumber.open('./data/Citi-2023-Annual-Report.pdf') as pdf:
    textlist = []

    for page in pdf.pages:
        text = page.extract_text()
        textlist.append(text)
        ##print("'"+text+"'")
        ##print("\n")
        ##print(type(text))
        ##print(text[-1]
'''